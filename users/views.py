from django.utils import timezone
import json

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.db import connection
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse, JsonResponse
import openpyxl
from .models import Channels, Client_Statut, Client_Target, Device, InternalUser, CustomUser, Produit, PromoDetails, PromoHeaders, Routes, Clients,PromoItemBasketHeaders,Client_Discounts, UserGroupe
from .forms import ChannelsForm, Client_TargetForm, DeviceForm, ProduitForm, PromoDetailsForm,PromoHeadersForm , PromotionSearchForm, NewPromotionForm, RouteForm, UserForm, AssignPromotionSearchForm, BasketForm, UserGroupeForm, client_discountsForm, client_statutForm, clientForm
from django.views.decorators.http import require_GET
from django.core import serializers
from django.db.utils import DatabaseError
from django.db.models import Q
from django.template.loader import render_to_string


import logging

logger = logging.getLogger(__name__)

def home(request):
    return render(request, 'home/home.html')

@login_required
def edit_route(request, route_id):
    if request.method == 'POST':
        route_description = request.POST.get('route_description')
        region_code = request.POST.get('region_description')

        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE Routes
                SET Route_Description = %s, Region_Code = %s
                WHERE Route_ID = %s
            """, [route_description, region_code, route_id])

        messages.success(request, 'Route updated successfully.')
        return redirect('routes')


def get_users(request):
    users = InternalUser.objects.all().values('UserCode', 'UserName', 'CityID')
    return JsonResponse({'users': list(users)})

# User login view for customers (authentication)
def user_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, 'Invalid username or password.')

    return render(request, 'authentication/login.html')


# User logout view
def user_logout(request):
    logout(request)
    return redirect('login')

#assign promo



def assign_promotions(request):
    promotions = PromoHeaders.objects.all()
    return render(request, 'promotions/assign_promotions.html', {'promotions': promotions})



def get_promo_client(request):
    promo_id = request.GET.get('promo_id')
    if not promo_id:
        return JsonResponse({'error': 'promo_id is required'}, status=400)

    try:
        promotion = PromoHeaders.objects.get(pk=promo_id)
    except PromoHeaders.DoesNotExist:
        return JsonResponse({'error': 'Promotion not found'}, status=404)

    assigned_clients = promotion.clients.all()
    unassigned_clients = Clients.objects.exclude(pk__in=assigned_clients)

    return JsonResponse({
        'unassigned_clients': list(unassigned_clients.values('Client_Code', 'Client_Description')),
        'assigned_clients': list(assigned_clients.values('Client_Code', 'Client_Description'))
    })
@csrf_exempt
def assign_client_to_promo(request):
    if request.method == 'POST':
        promo_id = request.POST.get('promo_id')
        client_code = request.POST.get('client_code')

        try:
            promotion = PromoHeaders.objects.get(pk=promo_id)
            client = Clients.objects.get(Client_Code=client_code)
        except (PromoHeaders.DoesNotExist, Clients.DoesNotExist):
            return JsonResponse({'error': 'Invalid promo or client'}, status=400)

        promotion.clients.add(client)
        return JsonResponse({'success': True})
    
@csrf_exempt
def remove_client_from_promo(request):
    if request.method == 'POST':
        promo_id = request.POST.get('promo_id')
        client_code = request.POST.get('client_code')

        try:
            promotion = PromoHeaders.objects.get(pk=promo_id)
            client = Clients.objects.get(Client_Code=client_code)
        except (PromoHeaders.DoesNotExist, Clients.DoesNotExist):
            return JsonResponse({'error': 'Invalid promo or client'}, status=400)

        promotion.clients.remove(client)
        return JsonResponse({'success': True})


def get_promo_user(request):
    promo_id = request.GET.get('promo_id')
    if promo_id:
        try:
            promo = PromoHeaders.objects.get(pk=promo_id)
            unassigned_users = InternalUser.objects.exclude(promotions=promo)
            assigned_users = InternalUser.objects.filter(promotions=promo)
            return JsonResponse({
                'unassigned_users': list(unassigned_users.values('UserCode', 'UserName')),
                'assigned_users': list(assigned_users.values('UserCode', 'UserName'))
            })
        except PromoHeaders.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Promotion not found'})
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)


@csrf_exempt
def assign_user_to_promo(request):
    if request.method == 'POST':
        promo_id = request.POST.get('promo_id')
        user_code = request.POST.get('user_code')

        try:
            promotion = PromoHeaders.objects.get(pk=promo_id)
            user = InternalUser.objects.get(UserCode=user_code)
        except (PromoHeaders.DoesNotExist, InternalUser.DoesNotExist):
            return JsonResponse({'error': 'Invalid promo or user'}, status=400)

        promotion.users.add(user)
        return JsonResponse({'success': True})
    
    
@csrf_exempt
def remove_user_from_promo(request):
    if request.method == 'POST':
        promo_id = request.POST.get('promo_id')
        user_code = request.POST.get('user_code')

        try:
            promotion = PromoHeaders.objects.get(pk=promo_id)
            user = InternalUser.objects.get(UserCode=user_code)
        except (PromoHeaders.DoesNotExist, InternalUser.DoesNotExist):
            return JsonResponse({'error': 'Invalid promo or user'}, status=400)

        promotion.users.remove(user)
        return JsonResponse({'success': True})

# basket + product
def define_basket(request):
    baskets = PromoItemBasketHeaders.objects.all()
    all_products = Produit.objects.all()
    form = ProduitForm()
    basket_form = BasketForm()
    context = {
        'baskets': baskets,
        'all_products': all_products,
        'form': form,
        'basket_form': basket_form

    }
    return render(request, 'promotions/define_basket.html', context)


def add_basket(request):
    if request.method == 'POST':
        form = BasketForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('define_basket')
    else:
        form = BasketForm()
    return render(request, 'promotions/define_basket.html', {'form': form})

def add_product(request):
    if request.method == "POST":
        form = ProduitForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('define_basket')  
    else:
        form = ProduitForm()
    return render(request, 'promotions/define_basket.html', {'form': form})

def view_products(request, basket_id):
    basket = get_object_or_404(PromoItemBasketHeaders, item_basket_id=basket_id)
    products = Produit.objects.all()
    if request.method == "POST":
        form = ProduitForm(request.POST)
        if form.is_valid():
            product = form.save()
            basket.products.add(product)
            return redirect('define_basket')
    else:
        form = ProduitForm()
    return render(request, 'view_products.html', {'basket': basket, 'products': products, 'form': form})


def add_product_to_basket(request, basket_id):
    basket = get_object_or_404(PromoItemBasketHeaders, item_basket_id=basket_id)
    if request.method == "POST":
        form = ProduitForm(request.POST)
        if form.is_valid():
            product = form.save()
            basket.products.add(product)
            return redirect('define_basket')
    else:
        form = ProduitForm()
    return redirect('define_basket')

def add_existing_product_to_basket(request, item_basket_id):
    #item_basket_id = request.POST.get('item_basket_id')
    basket = get_object_or_404(PromoItemBasketHeaders, item_basket_id=item_basket_id)
    if request.method == "POST":
        product_id = request.POST.get('existing_product')
        product = get_object_or_404(Produit, id=product_id)
        basket.products.add(product)
    return redirect('define_basket')

def delete_existing_product(request,item_basket_id , product_id):
    basket = get_object_or_404(PromoItemBasketHeaders, item_basket_id=item_basket_id)
    basket.products.remove(product_id)
   
    return JsonResponse({'message' : 'success'})

#promo
def promotions_list(request):
    promotions = PromoHeaders.objects.all()
    baskets = PromoItemBasketHeaders.objects.all()
    form = PromoHeadersForm()
    return render(request, 'promotions/home.html', {'promotions': promotions, 'form':form , 'baskets': baskets})

def promotions(request):
    if request.method == 'POST':
        form = PromoHeadersForm(request.POST)
        detail_form = PromoDetailsForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('promotions')
        else:
            print(form.errors)
    else:
        form = PromoHeadersForm()
        detail_form = PromoDetailsForm()
        baskets = PromoItemBasketHeaders.objects.all()

    context = {
        'form': form,
        'promotions': PromoHeaders.objects.all(),
        'detail_form': detail_form,
        'baskets': baskets
    }
    return render(request, 'promotions/home.html', context)

def create_promotion(request):
    baskets = PromoItemBasketHeaders.objects.all()

    if request.method == 'POST':
        form = PromoHeadersForm(request.POST)
        detail_form = PromoDetailsForm(request.POST)
        id = request.POST.get('basket')
        basket =PromoItemBasketHeaders.objects.get(item_basket_id = id)
        if form.is_valid() and detail_form.is_valid():
            promo = form.save()
            detail = detail_form.save(commit=False)
            detail.promotion_id = promo
            detail.basket = basket
            detail.save()
            
            messages.success(request, 'Promotion added successfully!')
            return redirect('promotions')
        else:
            messages.error(request, 'Error adding promotion.')
    else:
        form = PromoHeadersForm()
        detail_form = PromoDetailsForm()

    promotions = PromoHeaders.objects.all()
    return render(request, 'promotions/home.html', {'form': form, 'promotions': promotions, 'detail_form': detail_form,'baskets': baskets})


def edit_promotion(request, promotion_id):
    promotion = get_object_or_404(PromoHeaders, promotion_id=promotion_id)
    promo_detail = get_object_or_404(PromoDetails, promotion_id=promotion_id)
    baskets = PromoItemBasketHeaders.objects.all()

    if request.method == 'POST':
        form = PromoHeadersForm(request.POST, instance=promotion)
        detail_form = PromoDetailsForm(request.POST, instance=promo_detail)
        id = request.POST.get('basket2')
        basket =PromoItemBasketHeaders.objects.get(item_basket_id = id)
        if form.is_valid() and detail_form.is_valid():
            form.save()
            detail = detail_form.save(commit=False)
            detail.basket = basket
            detail.save()
            return redirect('promotions')
    else:
        form = PromoHeadersForm(instance=promotion)
        detail_form = PromoDetailsForm(instance=promo_detail)
    promotions = PromoHeaders.objects.all()
    return render(request, 'promotions/home.html', {
        'form': form,
        'detail_form': detail_form,
        'promotion': promotion,
        'promotions': promotions,
        'promo_detail': promo_detail,
        'baskets': baskets
    })



def delete_promotion(request, promotion_id):
    promotion = get_object_or_404(PromoHeaders, promotion_id=promotion_id)

    if request.method == 'POST':
        promotion.delete()
        return redirect('promotions')

    return render(request, 'promotions/home.html', {'promotion': promotion})

def get_promotion_data(request, promotion_id):
    promotion = get_object_or_404(PromoHeaders, promotion_id=promotion_id)
    promo_detail = get_object_or_404(PromoDetails, promotion_id=promotion_id)
    id = promo_detail.basket.item_basket_id
    data = {
        'promotion_id': promotion.promotion_id,
        'promotion_description': promotion.promotion_description,
        'promotion_type': promotion.promotion_type,
        'start_date': promotion.start_date,
        'end_date': promotion.end_date,
        'is_forced': promotion.is_forced,
        'is_active': promotion.is_active,
        'priority': promotion.priority,
        'promotion_apply': promotion.promotion_apply,
        'quantity_buy': promo_detail.quantity_buy,
        'types_buy': promo_detail.types_buy,
        'quantity_get': promo_detail.quantity_get,
        'types_get': promo_detail.types_get,
        'id':id
    }
    return JsonResponse(data)




def get_baskets(request):
    baskets = PromoItemBasketHeaders.objects.all()
    data = {
        'baskets': list(baskets.values('item_basket_id', 'item_basket_description'))
    }
    return JsonResponse(data)


def get_promotion_details(request, promotion_id):
    promo_details = PromoDetails.objects.filter(promotion_id=promotion_id)
    data = {
        'promo_details': list(promo_details.values('basket', 'quantity_buy', 'types_buy', 'quantity_get', 'types_get'))
    }
    return JsonResponse(data)





##### region
def get_regions():
    with connection.cursor() as cursor:
        cursor.execute("SELECT Region_Code, Region_Description FROM Regions")
        rows = cursor.fetchall()
    regions = [{'Region_Code': row[0], 'Region_Description': row[1]} for row in rows]
    return regions

def routes(request):
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT r.Route_ID, r.Route_Description, r.Region_Code, rg.Region_Description, r.RVSCode, r.RVSDescription
            FROM Routes r
            LEFT JOIN Regions rg ON r.Region_Code = rg.Region_Code
        """)
        routes_data = cursor.fetchall()

    routes_list = [{
        'Route_ID': row[0],
        'Route_Description': row[1],
        'Region_Code': row[2],
        'Region_Description': row[3],
        #'CreateBy': row[4],
        'HasClients': False,
        'RVSCode': row[4],
        'RVSDescription': row[5]
    } for row in routes_data]
    form = RouteForm()

    regions = get_regions()
    return render(request, 'routes/home.html', {'routes': routes_list, 'regions': regions, 'form': form})

def add_route(request):
    if request.method == 'POST':
        route_description = request.POST.get('route_description')
        region_code = request.POST.get('region_description')

        route = Routes(
            Branch_Code = '',
            Route_Description = route_description,
            Route_Alt_Description = '',
            Region_Code = region_code,
            
         )
        route.save()
        messages.success(request, 'Route added successfully.')
        return redirect('routes')

    regions = get_regions()
    return render(request, 'routes/home.html', {'regions': regions})


def edit_route(request, route_id):
    route = get_object_or_404(Routes, Route_ID=route_id)
    if request.method == "POST":
        form = RouteForm(request.POST, instance=route)
        if form.is_valid():
            form.save()
            return redirect('routes') 
    else:
        form = RouteForm(instance=route)
    return render(request, 'routes/home.html', {'form': form, 'route': route})

def delete_route(request, route_id):
    route = get_object_or_404(Routes, Route_ID=route_id)
    if request.method == 'POST':
        route.delete()
        return redirect('routes')
    return render(request, 'routes/home.html', {'route': route})


def get_routes_data(request, route_id):
    route= get_object_or_404(Routes, Route_ID=route_id)
    data = {
        'Branch_Code' :  route.Branch_Code,
        'Route_Description' : route.Route_Description,
        'Route_Alt_Description' : route.Route_Alt_Description,
        'Region_Code': route.Region_Code

    }
     
    return JsonResponse(data)


def affectation_clients_routes(request):
    if request.method == 'GET' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        action = request.GET.get('action')
        if action == 'get_routes':
            routes = Routes.objects.all().values('Route_ID', 'Route_Description')
            return JsonResponse(list(routes), safe=False)
        elif action == 'get_route_clients':
            route_id = request.GET.get('route_id')
            if not route_id:
                return JsonResponse({'error': 'No route ID provided'}, status=400)

            assigned_clients = Clients.objects.filter(Route_ID=route_id).values('Client_Code', 'Client_Description')
            available_clients = Clients.objects.filter(Route_ID=1).values('Client_Code', 'Client_Description')
            return JsonResponse({'assigned_clients': list(assigned_clients), 'available_clients': list(available_clients)})

    return render(request, 'routes/affectation_clients_routes.html')

def get_routes(request):
    routes = Routes.objects.filter(RVSCode__isnull=False).exclude(RVSCode='').values('Route_ID', 'Route_Description', 'RVSCode', 'RVSDescription')
    return JsonResponse({'routes': list(routes)})


def get_route_clients(request):
    route_id = request.GET.get('route_id')
    if route_id and route_id != '1':
        assigned_clients = Clients.objects.filter(Route_ID=route_id).values('Client_Code', 'Client_Description')
        available_clients = Clients.objects.filter(Route_ID=1).values('Client_Code', 'Client_Description')
    else:
        assigned_clients = []
        available_clients = Clients.objects.filter(Route_ID=1).values('Client_Code', 'Client_Description')

    return JsonResponse({'assigned_clients': list(assigned_clients), 'available_clients': list(available_clients)})




##################################

def assign_client_to_route(request):
    if request.method == 'POST':
        client_code = request.POST.get('client_code')
        route_id = request.POST.get('route_id')
        action = request.POST.get('action', 'assign')

        try:
            # Récupérer le client
            client = Clients.objects.get(Client_Code=client_code)
            
            if action == 'assign':
                # Récupérer l'instance de la route
                route_instance = Routes.objects.get(Route_ID=route_id)
                client.Route_ID = route_instance
            elif action == 'remove':
                # Optionnel : Vous pouvez définir Route_ID à None si vous utilisez une relation Nullable
                client.Route_ID = Routes.objects.get(Route_ID=1)  # Route_ID=1 pour aucune route

            client.save()
            return JsonResponse({'success': True})
        except Clients.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Client not found'})
        except Routes.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Route not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

#affct_
def affectation_routes_users(request):
    
    if request.method == 'GET' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        
        action = request.GET.get('action')
        if action == 'get_users':
            users = InternalUser.objects.all().values('UserCode', 'UserName', 'CityID')
            return JsonResponse(list(users), safe=False)
        elif action == 'get_user_routes':
            user_code = request.GET.get('user_code')
            if not user_code:
                return JsonResponse({'error': 'No user code provided'}, status=400)

            with connection.cursor() as cursor:
                cursor.execute("SELECT Route_ID, Route_Description FROM Routes WHERE RVSCode = '' ")
                unlinked_routes = cursor.fetchall()

                cursor.execute("SELECT Route_ID, Route_Description FROM Routes WHERE RVSCode = %s", [user_code])
                linked_routes = cursor.fetchall()

            unlinked_routes = [{'Route_ID': row[0], 'Route_Description': row[1]} for row in unlinked_routes]
            linked_routes = [{'Route_ID': row[0], 'Route_Description': row[1]} for row in linked_routes]

            return JsonResponse({'unlinked_routes': unlinked_routes, 'linked_routes': linked_routes})
        

    users = InternalUser.objects.all().values('UserCode', 'UserName', 'CityID')
    return render(request, 'routes/affectation_routes_users.html', {'users': users})

@csrf_exempt
def assign_user_to_route(request):
    if request.method == 'POST':
        route_id = request.POST.get('route_id')
        user_code = request.POST.get('user_code')
        action = request.POST.get('action')
        if not route_id or not user_code:
            return JsonResponse({'success': False, 'error': 'Invalid data'})
                    
        with connection.cursor() as cursor:
            if action == 'remove':
                cursor.execute("UPDATE Routes SET RVSCode = '' WHERE Route_ID = %s", [route_id])
            else:
                cursor.execute("SELECT Route_ID FROM Routes WHERE RVSCode = %s", [user_code])
                existing_route = cursor.fetchone()
                
                if existing_route:
                    return JsonResponse({'success': False, 'error': 'User is already assigned to a route'})
                cursor.execute("UPDATE Routes SET RVSCode = %s WHERE Route_ID = %s", [user_code, route_id])
                
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})



#device
def list_devices(request):
    devices = Device.objects.select_related('UserCode', 'Route_ID').all().values(
        'device_serial', 
        'device_name', 
        'UserCode__UserCode', 
        'device_status', 
        'type', 
        'Route_ID__Route_ID'
    )
    form = DeviceForm()
    return render(request, 'devices/home.html', {'devices': devices, 'form': form})

def add_device(request):
    if request.method == "POST":
        form = DeviceForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('devices')
    else:
        form = DeviceForm()
    return render(request, 'devices/home.html', {'form': form})


def get_device_data(request, device_serial):
    device = get_object_or_404(Device, device_serial=device_serial)
    data = {
        'device_serial': device.device_serial,
        'device_name': device.device_name,
        'UserCode': device.UserCode.UserCode if device.UserCode else None,
        'device_status': device.device_status,
        'type': device.type,
        'Route_ID': device.Route_ID.Route_ID if device.Route_ID else None
    }

    return JsonResponse(data)


def edit_device(request, device_serial):
    device = get_object_or_404(Device, device_serial=device_serial)
    if request.method == 'POST':
        form = DeviceForm(request.POST, instance=device)
        if form.is_valid():
            form.save()
            return redirect('devices')  
    else:
        form = DeviceForm(instance=device)
    return render(request, 'devices/home.html', {'form': form})

def remove_device_from_user(request, device_serial):
    device = get_object_or_404(Device, device_serial=device_serial)
    if request.method == 'POST':
        device.delete()
        return redirect('devices')
    return render(request, 'devices/home.html', {'device': device})

def home_device(request):
    devices = Device.objects.all()
    form = DeviceForm()
    return render(request, 'devices/home.html', {'devices': devices, 'form': form})



#client
def clients(request):
    if request.method == 'POST':
        form = clientForm(request.POST)
        if form.is_valid():
            form.save()
           
            return redirect('home_client')
    else:
        form = clientForm()
    
    return render(request, 'client/home_client.html', {'form': form})

def get_client_data(request, client_id):
    client = get_object_or_404(Clients, Client_Code=client_id)
    data = {
        'Client_Code': client.Client_Code,
        'Area_Code': client.Area_Code,
        'Client_Description': client.Client_Description,
        'Client_Alt_Description': client.Client_Alt_Description,
        'Payment_Term_Code': client.Payment_Term_Code,
        'Email': client.Email,
        'Address': client.Address,
        'Alt_Address': client.Alt_Address,
        'Contact_Person': client.Contact_Person,
        'Phone_Number': client.Phone_Number,
        'Barcode': client.Barcode,
        'Client_Status_ID': client.Client_Status_ID
    }

    return JsonResponse(data)

def edit_client(request, client_id):
    client = get_object_or_404(Clients, Client_Code=client_id)
    if request.method == "POST":
        form = clientForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            return redirect('home_client') 
    else:
        form = clientForm(instance=client)
    return render(request, 'client/home_client.html', {'form': form, 'client': client})

def home_client(request):
    clients = Clients.objects.all()
    clients_status= Client_Statut.objects.all()
    form = clientForm()
    return render(request, 'client/home_client.html', {'clients': clients, 'form': form,'clients_status':clients_status})

def delete_client(request, Client_Code):
    client = get_object_or_404(Clients, Client_Code=Client_Code)
    client.delete()
    return redirect('home_client')  

def search_client(request):
    clients = None
    query = ''
    if request.method == 'POST':
        query = request.POST.get('query', '')
        search_type = request.POST.get('search_type', 'client_code')
        if search_type == 'client_code':
            clients = Clients.objects.filter(Client_Code__icontains=query)
        elif search_type == 'description':
            clients = Clients.objects.filter(Client_Description__icontains=query)
        else:
            clients = Clients.objects.all()
    form = clientForm()
    return render(request, 'client/home_client.html', {'clients': clients, 'form': form, 'query': query})

#upload from excel 
def upload_excel(request):
    if request.method == 'POST' and request.FILES['xlxfile']:
        excel_file = request.FILES['xlxfile']
        
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=3, values_only=True):
            client_code = row[0]
            route = get_object_or_404(Routes, pk=row[12])

            if Clients.objects.filter(Client_Code=client_code).exists():
                messages.error(request, f"Client avec le code {client_code} existe déjà.")
            else:
                client = Clients(
                    Client_Code=client_code,
                    Area_Code=row[1],
                    Client_Description=row[2],
                    Client_Alt_Description=row[3],
                    Payment_Term_Code=row[4],
                    Email=row[5],
                    Address=row[6],
                    Alt_Address=row[7],
                    Contact_Person=row[8],
                    Phone_Number=row[9],
                    Barcode=row[10],
                    Client_Status_ID=row[11],
                    Route_ID=route
                )
                client.save()
                
                messages.success(request, f"Client {client.Client_Code} ajouté")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            print(row)
            
        return redirect('home_client')
    return render(request, 'client/upload_excel.html')


#statut client
def statut_client(request):
    if request.method == 'POST':
        form = client_statutForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('home_client_status')  
    else:
        form = client_statutForm()
        
    return render(request, 'client/status_client.html', {'form': form})

def get_client_statut_data(request, client_statut_id):
    client_statut = get_object_or_404(Client_Statut, Client_Statut_ID=client_statut_id)
    data = {
        'Client_Statut_ID': client_statut.Client_Statut_ID,
        'Statut_Description': client_statut.Statut_Description
    }
    return JsonResponse(data)

def edit_statut_client(request, client_statut_id):
    client = get_object_or_404(Client_Statut, Client_Statut_ID=client_statut_id)
    if request.method == "POST":
        form = client_statutForm(request.POST, instance=client)
        if form.is_valid():
            client.Client_Statut_ID = client_statut_id
            form.save()
            return redirect('home_client_status') 
    else:
        form = client_statutForm(instance=client)
    return render(request, 'client/status_client_modifier.html', {'form': form, 'client': client})


def home_client_status(request):
    clients = Client_Statut.objects.all()  
    form = client_statutForm()
    return render(request, 'client/home_client_status.html', {'client': clients, 'form': form})

#client discounts
def home_client_discounts(request):
    message = request.GET.get('message', '')
    form = client_discountsForm()
    clients = Client_Discounts.objects.all()
    return render(request, 'client/home_discounts.html', {'clients': clients, 'message': message, 'form' : form})



def get_client_Discount_data(request, client_Code):
    client_Discount = get_object_or_404(Client_Discounts, Client_Code=client_Code)
    data = {
        'Client_Code': client_Discount.Client_Code,
        'Trx_Code': client_Discount.Trx_Code,
        'Discounts': client_Discount.Discounts,
        'Month': client_Discount.Month,
        'Years': client_Discount.Years,
        'Discounts_label': client_Discount.Discounts_label,
        'Applied': client_Discount.Applied,
        'Stamp_Date': client_Discount.Stamp_Date,  
        'Affected_item_code': client_Discount.Affected_item_code,
    }
    return JsonResponse(data)



def client_discounts(request):
    error_message = None
    if request.method == "POST":
        client_code = request.POST.get("Client_Code")
        clients = Clients.objects.filter(Client_Code=client_code)
        
        if clients.exists():
            form = client_discountsForm(request.POST)
            if form.is_valid():
                form.save()
                return redirect('home_discounts')
            
        else:
            error_message = "Aucun client ne correspond à ce code client."
            form = client_discountsForm()
            return render(request, 'client/client_discounts.html', {'form': form, 'error_message': error_message})
    else:
        form = client_discountsForm()

    return render(request, 'client/client_discounts.html', {'form': form, 'error_message': error_message})


def edit_client_discounts(request, Client_Code):
    client = get_object_or_404(Client_Discounts, Client_Code=Client_Code)
    if request.method == "POST":
        form = client_discountsForm(request.POST, instance=client)
        if form.is_valid():
            client.Client_Code = Client_Code
            form.save()
            return redirect('home_discounts') 
    else:
        form = client_discountsForm(instance=client)
        form.fields['Client_Code'].widget.attrs['readonly'] = True
    return render(request, 'client/client_discounts_edit.html', {'form': form, 'client': client})

def delete_discounts(request, Client_Code):
    client = get_object_or_404(Client_Discounts, Client_Code=Client_Code)
    client.delete()
    return redirect('home_discounts') 


def delete_selected_clients(request):
    if request.method == 'POST':
        client_ids = request.POST.getlist('client_ids')
        if client_ids:
            Client_Discounts.objects.filter(Client_Code__in=client_ids).delete()
            messages.success(request, 'Selected clients have been deleted.')
        else:
            messages.error(request, 'No clients selected for deletion.')
    return redirect('home_discounts')
  
def search_client_discount(request):
    clients = None
    query = ''
    if request.method == 'POST':
        query = request.POST.get('query', '')
        search_type = request.POST.get('search_type', 'client_code')
        if search_type == 'client_code':
            clients = Client_Discounts.objects.filter(Client_Code=query)
        else:
            clients = Client_Discounts.objects.all()
    return render(request, 'client/home_discounts.html', {'clients': clients, 'query': query})
  
def upload_excel_discount(request):
    if request.method == 'POST' and request.FILES['xlxfile']:
        excel_file = request.FILES['xlxfile']
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=3, values_only=True):
            clients = Clients.objects.filter(Client_Code=row[0])
           
            if clients.exists():
                client = Client_Discounts(
                    Client_Code=row[0],
                    Trx_Code='',
                    Discounts=row[2],
                    Month=row[3],
                    Years=row[4],
                    Discounts_label=row[5],
                    Applied=0,
                    Stamp_Date=timezone.now(),
                    Affected_item_code='en instance'
                
                )
                client.save()  
                print(f"Client {client.Client_Code} ajouté")
            else:
                message = f"Les clients disponible sont tous ajoutés"

        for row in sheet.iter_rows(min_row=2, values_only=True):
            print(row)
        return redirect(f'/client/home_client_discount/?message={message}')     
    return render(request, 'client/upload_excel_discount.html')
    

#client target

def home_client_target(request):
    clients = Client_Target.objects.all() 
    form = Client_TargetForm()
    return render(request, 'client/home_client_target.html', {'clients': clients, 'form': form})

def get_target(request, Client_Code):
    client_t = get_object_or_404(Client_Target, Client_Code=Client_Code)
    data = {
        'Client_Code': client_t.Client_Code,
        'Month': client_t.Month,
        'years': client_t.years,
        'Targed_Achieved': client_t.Targed_Achieved,
        'Target_value': client_t.Target_value,

    }
    
    return JsonResponse(data)


def client_target(request):
    error_message = None
    if request.method == "POST":
        client_code = request.POST.get("Client_Code")
        clients = Clients.objects.filter(Client_Code=client_code)
        
        if clients.exists():
            form = Client_TargetForm(request.POST)
            if form.is_valid():
                form.save()
                return redirect('home_target')
            
        else:
            error_message = "Aucun client ne correspond à ce code client."
            form = Client_TargetForm()
            return render(request, 'client/client_target.html', {'form': form, 'error_message': error_message})
    else:
        form = Client_TargetForm()

    return render(request, 'client/client_target.html', {'form': form, 'error_message': error_message})


def edit_client_target(request, Client_Code):
    client = get_object_or_404(Client_Target, Client_Code=Client_Code)
    if request.method == "POST":
        form = Client_TargetForm(request.POST, instance=client)
        if form.is_valid():
            client.Client_Code = Client_Code
            form.save()
            return redirect('home_target') 
    else:
        form = Client_TargetForm(instance=client)
        form.fields['Client_Code'].widget.attrs['readonly'] = True
    return render(request, 'client/client_target_edit.html', {'form': form, 'client': client})

def delete_target(request, Client_Code):
    client = get_object_or_404(Client_Target, Client_Code=Client_Code)
    client.delete()
    return redirect('home_target') 


#channels
def channels(request):
    if request.method == 'POST':
        form = ChannelsForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('home_channel')
    else:
        form = ChannelsForm()
    
    channels_list = Channels.objects.all()  
    
    return render(request, 'client/channel.html', {'form': form, 'channels': channels_list})


def get_channel_data(request, channel_code):
    channel = get_object_or_404(Channels, channel_code=channel_code)
    data = {
        'channel_code': channel.channel_code,
        'Channel_description': channel.Channel_description
    }
    return JsonResponse(data)

def edit_channel(request, channel_code):
    client = get_object_or_404(Channels, channel_code=channel_code)
    if request.method == "POST":
        form = ChannelsForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            return redirect('home_channel') 
    else:
        form = ChannelsForm(instance=client)
    return render(request, 'client/modifier_channel.html', {'form': form, 'client': client})

def home_channel(request):
    clients = Channels.objects.all()
    form = ChannelsForm()
    return render(request, 'client/home_channel.html', {'form': form, 'clients': clients})

def delete_channel(request, channel_code):
    client = get_object_or_404(Channels, channel_code=channel_code)
    client.delete()
    return redirect('home_channel')  







#users


def users(request):
    users = InternalUser.objects.all()
    groupe = UserGroupe.objects.all()
    form = UserForm()
    return render(request, 'users/home.html', {'users': users, 'form': form, 'groupe': groupe})

def add_user(request):
    if request.method == "POST":
        form = UserForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('users')
    else:
        form = UserForm()
    return render(request, 'users/list_users.html', {'form': form})



def get_user_data(request, UserCode):
    user = get_object_or_404(InternalUser, UserCode=UserCode)
    data = {
        'UserCode': user.UserCode,
        'UserName': user.UserName,
        'PhoneNumber': user.PhoneNumber,
        'Grouping': user.Grouping,
        'IsBlocked': user.IsBlocked,
        'LoginName': user.LoginName,
        'AreaCode': user.AreaCode,
        'CityID': user.CityID,
        'RouteCode': user.RouteCode,
        'ParentCode': user.ParentCode
    }

    return JsonResponse(data)






def edit_user(request, UserCode):
    user = get_object_or_404(InternalUser, UserCode=UserCode)
    if request.method == "POST":
        form = UserForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            print(user.IsBlocked)  
            return redirect('users')
    else:
        form = UserForm(instance=user)
    
    return render(request, 'users/home.html', {'form': form, 'user': user})

def delete_user(request, UserCode):
    user = get_object_or_404(InternalUser, UserCode=UserCode)
    user.delete()
    return redirect('users') 


def delete_groupe(request, UserCode):
    user = get_object_or_404(UserGroupe, UserCode=UserCode)
    if request.method == 'POST':
        user.delete()
        return redirect('users')
    return render(request, 'users/home.html', {'user': user})

def search_user(request):
    users = None
    query = ''
    if request.method == 'POST':
        query = request.POST.get('query', '')
        search_type = request.POST.get('search_type', 'UserCode')
        if search_type == 'UserCode':
            users = InternalUser.objects.filter(UserCode__icontains=query)
        elif search_type == 'UserName':
            users = InternalUser.objects.filter(UserName__icontains=query)
        elif search_type == 'PhoneNumber':
            users = InternalUser.objects.filter(PhoneNumber=query)
        elif search_type == 'CityID':
            users = InternalUser.objects.filter(CityID=query)
        else:
            users = InternalUser.objects.all()
    form = UserForm()
    return render(request, 'users/home.html', {'users': users, 'form': form, 'query': query})


#groupe

def home_groupe(request):
    groupes = UserGroupe.objects.all()
    return render(request, 'users/Groupes.html', {'groupes': groupes})

def add_groupe(request):
    if request.method == 'POST':
        form = UserGroupeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('home_groupe')
    else:
        form = UserGroupeForm()
    return render(request, 'users/Groupes.html', {'form': form})

def edit_groupe(request, Code_groupe):
    groupe = get_object_or_404(UserGroupe, Code_groupe=Code_groupe)
    if request.method == 'POST':
        form = UserGroupeForm(request.POST, instance=groupe)
        if form.is_valid():
            form.save()
            return redirect('home_groupe')
    else:
        form = UserGroupeForm(instance=groupe)
    return render(request, 'users/Groupes.html', {'form': form, 'groupe': groupe})

def delete_groupe(request, Code_groupe):
    groupe = get_object_or_404(UserGroupe, Code_groupe=Code_groupe)
    if request.method == 'POST':
        groupe.delete()
        return redirect('home_groupe')
    return render(request, 'users/Groupes.html', {'groupe': groupe})

def get_group_data(request, Code_groupe):
    group = get_object_or_404(UserGroupe, Code_groupe=Code_groupe)
    data = {
        'Code_groupe': group.Code_groupe,
        'Groupe_description': group.Groupe_description,
    }

    return JsonResponse(data)




def search_promotions(request):
    if request.method == 'GET':
        search_query = request.GET.get('query', '')
        promotions = PromoHeaders.objects.filter(promotion_description__icontains=search_query)
        promo_list = [
            {
                'promotion_id': promo.promotion_id,
                'promotion_description': promo.promotion_description,
                'start_date': promo.start_date,
                'end_date': promo.end_date,
            }
            for promo in promotions
        ]
        return JsonResponse({'promotions': promo_list})



